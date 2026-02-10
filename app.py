import streamlit as st
import pandas as pd
import re
import os
import json
from io import BytesIO
import zipfile
from datetime import datetime
import google.generativeai as genai
import xlrd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

load_dotenv()

# Page configuration
st.set_page_config(
    page_title="Quote Processor",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
    <style>
    .main { background: linear-gradient(to bottom right, #EFF6FF, #E0E7FF); }
    .stButton>button { width: 100%; background-color: #4F46E5; color: white; font-weight: bold; padding: 0.75rem; border-radius: 0.5rem; }
    .stButton>button:hover { background-color: #4338CA; }
    </style>
""", unsafe_allow_html=True)

# ============================================================================
# VERTICAL QUOTES PROCESSOR CLASS
# ============================================================================

class VerticalQuoteProcessor:
    """Handles processing of vertical format Excel quote files"""
    
    def __init__(self, density=7.85e-6):
        self.density = density
    
    def find_matching_sheet(self, excel_file):
        """Find the appropriate sheet to process in the Excel file"""
        sheet_names = excel_file.sheet_names
        matching_sheets = [
            sheet for sheet in sheet_names
            if (("part" in sheet.lower() and "cost" in sheet.lower()) or 
                ("common" in sheet.lower()) or 
                ("part" in sheet.lower()) or 
                sheet.isdigit())
            and ("deleted" not in sheet.lower() and "other" not in sheet.lower())
        ]
        return matching_sheets[0] if matching_sheets else None
    
    def find_right_of(self, label, df, offset=1, start_row=0, end_row=None, regex=False):
        """Find value to the right of a label in the dataframe"""
        if end_row is None:
            end_row = len(df)

        if regex:
            pattern = re.compile(label, re.IGNORECASE)
        else:
            label_norm = re.sub(r'\s+', ' ', label.replace('\xa0', ' ').strip().lower())

        for r in range(start_row, end_row):
            for c in range(len(df.columns)):
                cell = str(df.iat[r, c]).strip()
                cell_norm = re.sub(r'\s+', ' ', cell.lower())

                match = (
                    re.search(pattern, cell_norm) if regex
                    else label_norm in cell_norm
                )

                if match:
                    for i in range(c + offset, len(df.columns)):
                        val = df.iat[r, i]
                        if pd.notna(val) and str(val).strip() != "":
                            return str(val).strip()
        return None
    
    def extract_validity_mapping(self, df):
        """Extract validity values mapped to part numbers"""
        validity_map = {}
        val_col = None
        val_header_row = None

        for r in range(min(5, len(df))):
            for c in range(len(df.columns)):
                cell = str(df.iat[r, c]).strip().lower()
                if "unit cost" not in cell and re.search(r"\bval\b", cell):
                    val_col = c
                    val_header_row = r
                    break
            if val_col is not None:
                break

        if val_col is None:
            return validity_map

        for r in range(val_header_row + 1, len(df)):
            try:
                val_cell = df.iat[r, val_col]

                if pd.notna(val_cell):
                    val_str = str(val_cell).strip()

                    if val_str and val_str.replace('.', '', 1).replace('-', '', 1).isdigit():
                        part_no = None

                        for c in range(val_col):
                            cell = str(df.iat[r, c]).strip()

                            if re.match(r'^[A-Z0-9][A-Z0-9\-]{5,}$', cell, re.IGNORECASE):
                                part_no = cell.upper()
                                break

                        if part_no:
                            validity_map[part_no] = val_str

            except (IndexError, ValueError):
                continue

        return validity_map
    
    def is_pipe_or_tube(self, part_name):
        """Check if part is a pipe or tube"""
        if not part_name:
            return False
        part_name_lower = str(part_name).lower()
        return bool(re.search(r'\b(pipe|pipes|tube|tubes|tubing)\b', part_name_lower))
    
    def looks_like_material_grade(self, text):
        """Check if text looks like a material grade"""
        if not text or len(text) > 15:
            return False

        text = str(text).strip().upper()

        grade_patterns = [
            r'^[A-Z]{1,4}\d+[A-Z]?$',
            r'^(CR|HR|EN|IS|ASTM|SAE|DIN|JIS)[A-Z0-9]+$',
            r'^[A-Z0-9]{2,8}$',
        ]

        for pattern in grade_patterns:
            if re.match(pattern, text):
                return True

        return False
    
    def safe_float(self, val, default=0.0):
        """Safely convert value to float"""
        try:
            if val is None:
                return default
            return float(str(val).replace(',', '').strip())
        except:
            return default

    def safe_int(self, val, default=1):
        """Safely convert value to int"""
        try:
            if val is None:
                return default
            return int(float(str(val).replace(',', '').strip()))
        except:
            return default
    
    def extract_assembly_info(self, df):
        """Extract assembly part number, name, and MOD"""
        assembly_part_no = self.find_right_of(r"\bassy\s*part\s*no\.?\s*[:\-]+", df, 1, start_row=0, end_row=5, regex=True)
        assembly_part_name = self.find_right_of(r"\bassy\s*part\s*name\.?\s*[:\-]+", df, 1, start_row=0, end_row=5, regex=True)
        assembly_mod = self.find_right_of(r"\bmod\.?\s*[:\-]*", df, offset=1, start_row=0, end_row=5, regex=True)
        
        if assembly_mod is None:
            assembly_mod = "-"
        
        return assembly_part_no, assembly_part_name, assembly_mod
    
    def extract_part_info(self, df, start_row, end_row):
        """Extract part number and name from a section"""
        part_no = None
        part_name = None
        part_no_row = None
        part_no_col = None

        for r in range(start_row, min(start_row + 3, end_row)):
            for c in range(len(df.columns)):
                try:
                    cell = str(df.iat[r, c]).strip()
                    if re.search(r"\bpart\s*no\.?\s*[:\-]+", cell, re.IGNORECASE):
                        if not re.search(r"\bassy", cell, re.IGNORECASE):
                            part_no_row = r
                            part_no_col = c

                            for offset in range(1, 5):
                                if c + offset < len(df.columns):
                                    candidate = str(df.iat[r, c + offset]).strip()
                                    if candidate and candidate.lower() not in ['nan', '', '-', ':']:
                                        if re.match(r'^[A-Z0-9][A-Z0-9\-]{5,}$', candidate, re.IGNORECASE):
                                            part_no = candidate
                                            part_no_col = c + offset

                                            # Find part name
                                            for name_offset in range(1, 5):
                                                name_col = part_no_col + name_offset
                                                if name_col >= len(df.columns):
                                                    break

                                                name_candidate = str(df.iat[r, name_col]).strip()

                                                if not name_candidate or name_candidate.lower() in ['nan', '', '-', ':', 'none']:
                                                    continue

                                                if self.looks_like_material_grade(name_candidate):
                                                    break

                                                name_lower = name_candidate.lower()
                                                exact_skip_keywords = ['raw', 'material', 'mod', 'rev', 'rm', 'spec', 'grade', 'mod:', 'rev:', 'mod-', 'mod.']
                                                if name_lower in exact_skip_keywords:
                                                    continue

                                                if any(phrase in name_lower for phrase in ['raw material', 'rm spec', 'material grade']):
                                                    break

                                                if re.match(r'^[\d\.\,\-\/\s]+$', name_candidate):
                                                    continue

                                                if len(name_candidate) >= 3 and re.search(r'[A-Za-z]', name_candidate):
                                                    if not (len(name_candidate) <= 6 and re.match(r'^[A-Z]{1,3}\d+[A-Z]?$', name_candidate.upper())):
                                                        part_name = name_candidate
                                                        break

                                            break
                            break
                except:
                    continue

            if part_no:
                break
        
        # If part name not found, try alternative search
        if part_no and part_no_row is not None and part_no_col is not None and part_name is None:
            material_grade_start_col = None
            for c in range(part_no_col + 1, len(df.columns)):
                cell = str(df.iat[part_no_row, c]).strip().lower()
                if re.search(r'(raw\s*material|rm\s*spec|material\s*grade|grade)', cell):
                    material_grade_start_col = c
                    break

            end_search = material_grade_start_col if material_grade_start_col else min(part_no_col + 15, len(df.columns))

            for c in range(part_no_col + 1, end_search):
                try:
                    name_candidate = str(df.iat[part_no_row, c]).strip()

                    if not name_candidate or name_candidate.lower() in ['nan', '', '-', ':', 'none']:
                        continue

                    if self.looks_like_material_grade(name_candidate):
                        break

                    name_lower = name_candidate.lower()
                    skip_keywords = ['raw', 'material', 'mod', 'rev', 'rm', 'spec', 'grade', 'unit', 'cost']
                    if any(kw == name_lower for kw in skip_keywords):
                        continue

                    if re.match(r'^[\d\.\,\-\/\s]+$', name_candidate):
                        continue

                    if len(name_candidate) >= 3 and re.search(r'[A-Za-z]', name_candidate):
                        if not (len(name_candidate) <= 6 and re.match(r'^[A-Z]{1,3}\d+[A-Z]?$', name_candidate.upper())):
                            part_name = name_candidate
                            break
                except:
                    continue

        if part_name is None:
            part_name = ""
        
        return part_no, part_name, part_no_row, part_no_col
    
    def extract_mod(self, df, part_no_row, end_row, assembly_part_no, assembly_mod, part_no):
        """Extract MOD/Rev information"""
        rev = None
        if part_no_row is not None:
            search_range = min(part_no_row + 15, end_row)

            for r in range(part_no_row, search_range):
                if rev is not None:
                    break

                for c in range(len(df.columns)):
                    try:
                        cell = str(df.iat[r, c]).strip()
                        cell_lower = cell.lower()

                        is_mod_label = (re.search(r'\bmod\b', cell_lower) or
                                       re.search(r'\brev\b', cell_lower) or
                                       cell_lower in ['mod', 'rev', 'mod:', 'rev:', 'mod -', 'mod-'])

                        if is_mod_label:
                            for i in range(c + 1, min(c + 5, len(df.columns))):
                                try:
                                    val = df.iat[r, i]
                                    val_str = str(val).strip()

                                    if not val_str or val_str.lower() in ['nan', '', '-']:
                                        continue

                                    if pd.notna(val) and val_str.lower() not in ['mod', 'rev']:
                                        rev = val_str
                                        break
                                except:
                                    continue

                            if rev:
                                break
                    except:
                        continue

        if rev is None:
            if part_no and assembly_part_no:
                part_no_normalized = str(part_no).strip().upper()
                assembly_part_no_upper = str(assembly_part_no).strip().upper()
                if part_no_normalized == assembly_part_no_upper:
                    rev = assembly_mod if assembly_mod else "-"
                else:
                    rev = "-"
            else:
                rev = "-"
        
        return rev
    
    def extract_material_grade(self, df, start_row, end_row):
        """Extract material grade"""
        material_grade = None

        patterns = [
            r"\braw\s*material\s*specs?\.?\s*[:\-]*",
            r"\brm\s*specs?\.?\s*[:\-]*",
            r"\bmaterial\s*grade\.?\s*[:\-]*",
            r"\bgrade\.?\s*[:\-]*"
        ]

        for pattern in patterns:
            if material_grade is not None:
                break
            for row_idx in range(start_row, end_row):
                row = df.iloc[row_idx]
                for col_idx, cell in enumerate(row):
                    if re.search(pattern, str(cell), re.IGNORECASE):
                        for offset in range(1, 4):
                            if col_idx + offset < len(row):
                                candidate = str(row.iloc[col_idx + offset]).strip()
                                if candidate and candidate.lower() not in ['nan', '', '-', ':']:
                                    if re.match(r"^(?=.*[A-Za-z])[A-Za-z0-9./-]+$", candidate):
                                        material_grade = candidate
                                        break
                        if material_grade:
                            break
                if material_grade:
                    break
        
        return material_grade
    
    def extract_dimensions_and_weights(self, df, start_row, end_row):
        """Extract all dimension and weight data"""
        thickness = self.safe_float(self.find_right_of("Full Sheet size", df, 1, start_row, end_row) or 0)
        sheet_width = self.safe_float(self.find_right_of("Full Sheet size", df, 2, start_row, end_row) or 0)
        sheet_length = self.safe_float(self.find_right_of("Full Sheet size", df, 3, start_row, end_row) or 0)
        no_of_components = self.safe_int(self.find_right_of("Full Sheet size", df, 6, start_row, end_row) or 1)
        rate = self.safe_float(self.find_right_of("Full Sheet size", df, 7, start_row, end_row) or 0)

        blank_width = self.safe_float(self.find_right_of("Shear Size", df, 2, start_row, end_row) or 0)
        blank_length = self.safe_float(self.find_right_of("Shear Size", df, 3, start_row, end_row) or 0)
        no_of_components_blank = self.safe_int(self.find_right_of("Shear Size", df, 6, start_row, end_row) or 1)

        finish_weight_kg = self.extract_finish_weight(df, start_row, end_row)
        scrap_rate = self.safe_float(self.find_right_of("Scrap", df, 7, start_row, end_row) or 0)
        
        return {
            'thickness': thickness,
            'sheet_width': sheet_width,
            'sheet_length': sheet_length,
            'no_of_components': no_of_components,
            'rate': rate,
            'blank_width': blank_width,
            'blank_length': blank_length,
            'no_of_components_blank': no_of_components_blank,
            'finish_weight_kg': finish_weight_kg,
            'scrap_rate': scrap_rate
        }
    
    def extract_finish_weight(self, df, start_row, end_row):
        """Extract finished weight"""
        finish_weight_kg = 0.0
        finished_wt_row = None

        for r in range(start_row, end_row):
            for c in range(len(df.columns)):
                cell = str(df.iat[r, c]).strip().lower()
                if "finished" in cell and "wt" in cell:
                    finished_wt_row = r
                    break
            if finished_wt_row is not None:
                break

        if finished_wt_row is not None:
            wt_pc_col = None
            search_start = max(start_row, finished_wt_row - 5)

            for r in range(search_start, finished_wt_row):
                for c in range(len(df.columns)):
                    cell = str(df.iat[r, c]).strip().lower()
                    if re.search(r'wt\.?\s*\/\s*pc', cell):
                        wt_pc_col = c
                        break
                if wt_pc_col is not None:
                    break

            if wt_pc_col is not None:
                try:
                    val = df.iat[finished_wt_row, wt_pc_col]
                    if pd.notna(val):
                        finish_weight_kg = self.safe_float(val)
                except:
                    pass

            if finish_weight_kg == 0.0:
                for c in range(len(df.columns) - 1, -1, -1):
                    try:
                        val = df.iat[finished_wt_row, c]
                        if pd.notna(val) and str(val).strip() != "":
                            cell_str = str(val).strip().lower()
                            if "finished" not in cell_str and "wt" not in cell_str:
                                finish_weight_kg = self.safe_float(val)
                                break
                    except:
                        continue
        
        return finish_weight_kg
    
    def process_quote(self, df, output_path):
        """Main processing function for vertical quotes"""
        assembly_part_no, assembly_part_name, assembly_mod = self.extract_assembly_info(df)
        assembly_validity = 1

        validity_map = self.extract_validity_mapping(df)

        assembly_part_no_upper = str(assembly_part_no).strip().upper() if assembly_part_no else None

        assembly_matches_component = False
        if assembly_part_no_upper and assembly_part_no_upper in validity_map:
            assembly_matches_component = True

        output_rows = []

        if not assembly_matches_component:
            output_rows.append({
                "LEVEL": "",
                "PART NO": assembly_part_no,
                "PART NAME": assembly_part_name,
                "Type": "",
                "MOD": assembly_mod,
                "Commodity": "",
                "Validity": assembly_validity,
                "MATERIAL GRADE": "",
                "THK MM": "",
                "BLANK WIDTH (W)": "",
                "BLANK LENGTH (L)": "",
                "NO OF COMPONENTS/BLANK": "",
                "SHEET WIDTH MM": "",
                "SHEET LENGTH MM": "",
                "NO OF COMPONENTS": "",
                "FINISH WEIGHT IN KG": "",
                "RATE": "",
                "SCRAP RATE": "",
            })

        # Locate each "Part No" section
        part_indices = []
        for r in range(len(df)):
            for c in range(len(df.columns)):
                val = str(df.iat[r, c]).strip()
                if re.search(r"\bpart\s*no\.?\s*[:\-]+", val, re.IGNORECASE):
                    if not re.search(r"\bassy", val, re.IGNORECASE):
                        part_indices.append(r)

        part_indices.append(len(df))

        # Loop through each part section
        for idx in range(len(part_indices) - 1):
            start_row = part_indices[idx]
            end_row = part_indices[idx + 1]

            part_no, part_name, part_no_row, part_no_col = self.extract_part_info(df, start_row, end_row)
            
            if not part_no:
                continue
            
            rev = self.extract_mod(df, part_no_row, end_row, assembly_part_no_upper, assembly_mod, part_no)
            material_grade = self.extract_material_grade(df, start_row, end_row)
            dimensions = self.extract_dimensions_and_weights(df, start_row, end_row)

            validity = 1
            if part_no:
                part_no_upper = str(part_no).strip().upper()
                if part_no_upper in validity_map:
                    val_str = validity_map[part_no_upper]
                    try:
                        if '.' in val_str:
                            validity = float(val_str)
                        else:
                            validity = int(val_str)
                    except:
                        validity = val_str

            if dimensions['no_of_components'] == 0:
                dimensions['no_of_components'] = 1

            output_rows.append({
                "LEVEL": "",
                "PART NO": part_no,
                "PART NAME": part_name,
                "Type": "",
                "MOD": rev,
                "Commodity": "",
                "Validity": validity,
                "MATERIAL GRADE": material_grade,
                "THK MM": dimensions['thickness'],
                "BLANK WIDTH (W)": dimensions['blank_width'],
                "BLANK LENGTH (L)": dimensions['blank_length'],
                "NO OF COMPONENTS/BLANK": dimensions['no_of_components_blank'],
                "SHEET WIDTH MM": dimensions['sheet_width'],
                "SHEET LENGTH MM": dimensions['sheet_length'],
                "NO OF COMPONENTS": dimensions['no_of_components'],
                "FINISH WEIGHT IN KG": dimensions['finish_weight_kg'],
                "RATE": dimensions['rate'],
                "SCRAP RATE": dimensions['scrap_rate'],
            })

        output_df = pd.DataFrame(output_rows)

        import xlsxwriter
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            output_df.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            yellow_format = workbook.add_format({
                'bg_color': '#FFFF00',
                'border': 1
            })

            columns_to_highlight = ['LEVEL', 'Type', 'Commodity']

            for col_name in columns_to_highlight:
                if col_name in output_df.columns:
                    col_idx = output_df.columns.get_loc(col_name)

                    for row_num in range(1, len(output_df) + 1):
                        worksheet.write(row_num, col_idx, output_df.iloc[row_num - 1][col_name], yellow_format)

        return output_df
    
    def process_file(self, uploaded_file):
        """Process a single uploaded file"""
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            sheet = self.find_matching_sheet(excel_file)
            
            if not sheet:
                return {'status': 'skipped', 'filename': uploaded_file.name, 'reason': 'No matching sheet'}
            
            df = pd.read_excel(uploaded_file, sheet_name=sheet, header=None)
            output_buffer = BytesIO()
            self.process_quote(df, output_buffer)
            output_buffer.seek(0)
            
            return {
                'status': 'success',
                'filename': uploaded_file.name,
                'output_filename': os.path.splitext(uploaded_file.name)[0] + "_Processed.xlsx",
                'buffer': output_buffer
            }
        except Exception as e:
            return {'status': 'error', 'filename': uploaded_file.name, 'reason': str(e)}


# ============================================================================
# HORIZONTAL QUOTES PROCESSOR CLASS
# ============================================================================

class HorizontalQuoteProcessor:
    """Handles processing of horizontal format Excel quote files using AI"""
    
    REQUIRED_COLUMNS = {
        'Part No': 'PART NO',
        'Part Name': 'PART NAME',
        'Mod': 'MOD',
        'Val': 'Validity',
        'Material': 'MATERIAL GRADE',
        'Thk mm': 'THK MM',
        'Blank Width mm': 'BLANK WIDTH (W)',
        'Blank Length mm': 'BLANK LENGTH (L)',
        'No Of Components per blank': 'NO OF COMPONENTS/BLANK',
        'Sheet Width mm': 'SHEET WIDTH MM',
        'Sheet Length mm': 'SHEET LENGTH MM',
        'No Of Components per sheet': 'NO OF COMPONENTS',
        'Fin Weight (Physically Checked)': 'FINISH WEIGHT IN KG'
    }

    CORE_REQUIRED = ['Part No', 'Part Name']
    ADDITIONAL_REQUIRED = ['Mod', 'Val']
    
    def __init__(self, api_key):
        self.api_key = api_key
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-1.5-flash')
    
    def identify_columns_with_ai(self, sheet_name, columns, sample_rows):
        """Use Gemini AI to identify columns"""
        context_table = []
        for i, header in enumerate(columns):
            samples = []
            for row in sample_rows:
                if i < len(row) and row[i] not in [None, ""]:
                    sample_value = str(row[i]).strip()
                    if sample_value and sample_value.lower() != "nan":
                        samples.append(sample_value)
            sample_str = ", ".join(samples[:3]) if samples else "(no data)"
            context_table.append(f"'{header}' → [{sample_str}]")

        prompt = f"""
                You are analyzing an Excel sheet to identify columns for quote/part data extraction.

                SHEET NAME: {sheet_name}

                AVAILABLE COLUMNS WITH SAMPLE DATA:
                {chr(10).join(context_table)}

                REQUIRED COLUMNS TO IDENTIFY:
                {json.dumps(list(self.REQUIRED_COLUMNS.keys()), indent=2)}

                IDENTIFICATION RULES:
                1. Part No: Contains alphanumeric codes (e.g., 554729500101) - MUST be long codes (10+ characters)
                2. Part Name: Descriptive text (e.g., "BRACKET ASSY BOOSTER MTG")
                3. Mod: Short codes (e.g., "b", "nr", "A") - May also be labeled as "Revision", "Rev No", "MOD"
                4. Val/Validity: Numbers representing validity period (e.g., 1, 12, 24)
                5. Material: Material grades (e.g., "E34", "Steel", "CR4")
                6. Thk mm: Thickness in millimeters (numeric values like 2.5, 1.8)
                7. Blank Width mm: Width measurements in mm (numeric)
                8. Blank Length mm: Length measurements in mm (numeric)
                9. No Of Components per blank: Count of components (numeric)
                10. Sheet Width mm: Sheet width in mm (numeric)
                11. Sheet Length mm: Sheet length in mm (numeric)
                12. No Of Components per sheet: Count per sheet (numeric)
                13. Fin Weight: FINISH weight - Look for "Fin Weight", "Finish Weight", "Final Weight"
                - DO NOT select: "Input Weight", "I/P Weight", "Weight/Car"

                Return ONLY valid JSON (no markdown):
                {{
                    "is_relevant_sheet": true,
                    "found_columns": {{
                        "Part No": "<exact_header_or_null>",
                        "Part Name": "<exact_header_or_null>",
                        "Mod": "<exact_header_or_null>",
                        "Val": "<exact_header_or_null>",
                        "Material": "<exact_header_or_null>",
                        "Thk mm": "<exact_header_or_null>",
                        "Blank Width mm": "<exact_header_or_null>",
                        "Blank Length mm": "<exact_header_or_null>",
                        "No Of Components per blank": "<exact_header_or_null>",
                        "Sheet Width mm": "<exact_header_or_null>",
                        "Sheet Length mm": "<exact_header_or_null>",
                        "No Of Components per sheet": "<exact_header_or_null>",
                        "Fin Weight (Physically Checked)": "<exact_header_or_null>"
                    }},
                    "confidence": "high"
                }}
                """

        try:
            response = self.model.generate_content(prompt)
            response_text = response.text.strip()

            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.startswith('```'):
                response_text = response_text[3:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            response_text = response_text.strip()

            return json.loads(response_text)
        except Exception as e:
            return None
    
    def fuzzy_match_columns(self,columns):
        """
        Fallback method: Match columns using string similarity
        """
        mapping = {}

        # Define variations for each required column
        variations = {
            'Part No': ['part no', 'part number', 'partno', 'part_no', 'part-no'],
            'Loose Part No': ['loose part no', 'loose part number', 'child part no', 'sub part no'],
            'Part Name': ['part name', 'part description', 'description', 'partname'],
            'Mod': ['mod', 'model', 'modification', 'revision', 'rev no', 'rev', 'rev.'],
            'Val': ['val', 'validity', 'valid'],
            'Material': ['material', 'material grade', 'grade', 'mat'],
            'Thk mm': ['thk mm', 'thickness', 'thk', 'thickness mm'],
            'Blank Width mm': ['blank width', 'blank width mm', 'width', 'blank_width'],
            'Blank Length mm': ['blank length', 'blank length mm', 'length', 'blank_length'],
            'No Of Components per blank': ['components per blank', 'no of components per blank', 'comp/blank', 'compo per blank', 'no of components/blank', 'components/blank'],
            'Sheet Width mm': ['sheet width', 'sheet width mm', 'sht width'],
            'Sheet Length mm': ['sheet length', 'sheet length mm', 'sht length'],
            'No Of Components per sheet': ['components per sheet', 'no of components per sheet', 'comp/sheet', 'compo per sheet', 'no of components/sheet', 'components/sheet', 'no of components'],
            'Fin Weight (Physically Checked)': ['fin weight', 'finish weight', 'final weight', 'physically checked']
        }

        weight_exclusions = ['input', 'i/p', 'inp', 'car', 'blank', 'raw', 'initial']

        for required_col, var_list in variations.items():
            for col in columns:
                col_lower = str(col).lower().strip()
                col_normalized = col_lower.replace('/', '').replace('-', '').replace('_', ' ').replace('.', '')

                if required_col == 'Fin Weight (Physically Checked)':
                    if any(excl in col_lower for excl in weight_exclusions):
                        continue
                    if any(var in col_lower for var in var_list):
                        mapping[required_col] = col
                        break
                else:
                    for variation in var_list:
                        var_normalized = variation.replace('/', '').replace('-', '').replace('_', ' ')

                        if var_normalized in col_normalized:
                            mapping[required_col] = col
                            break

                    if required_col in mapping:
                        break

        return mapping
    
    def find_header_row(self, file_bytes, sheet_name, file_ext):
        """Find the row containing headers"""
        try:
            engine = 'xlrd' if file_ext == '.xls' else 'openpyxl'
            df_preview = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=20, engine=engine)
            
            header_keywords = ['part', 'no', 'name', 'mod', 'material', 'val', 'validity', 'thk', 'thickness', 'width', 'length', 'component', 'weight']
            
            for i in range(min(20, len(df_preview))):
                row = df_preview.iloc[i]
                if row.notna().sum() < 3:
                    continue
                keyword_count = sum(1 for val in row if pd.notna(val) and any(kw in str(val).lower() for kw in header_keywords))
                if keyword_count >= 5:
                    return i
            
            return 0
        except:
            return 0
    
    def is_valid_part_no(self, part_no):
        """Validate part number format"""
        part_str = str(part_no).strip()
        if len(part_str.replace(' ', '')) < 10:
            return False
        if sum(c.isdigit() for c in part_str) < 8:
            return False
        if any(kw in part_str.lower() for kw in ['part no', 'part number', 'welding']):
            return False
        return True
    
    def has_sufficient_data(self, row):
        """Check if row has enough data"""
        return sum(1 for val in row if pd.notna(val) and (not isinstance(val, str) or val.strip() != '')) > 4
    
    def process_sheet(self, df, sheet_name):
        """Process a single sheet"""
        sample_rows = [df.iloc[idx].tolist() for idx in range(min(5, len(df)))]
        
        ai_result = self.identify_columns_with_ai(sheet_name, df.columns.tolist(), sample_rows)
        
        if ai_result and not ai_result.get('is_relevant_sheet', True):
            return None
        
        if ai_result and ai_result.get('found_columns'):
            column_mapping = {k: v for k, v in ai_result['found_columns'].items() if v is not None}
        else:
            column_mapping = self.fuzzy_match_columns(df.columns)
        
        has_core = all(col in column_mapping and column_mapping[col] is not None for col in self.CORE_REQUIRED)
        if not has_core:
            return None
        
        additional_found = sum(1 for col in self.ADDITIONAL_REQUIRED if col in column_mapping and column_mapping[col] is not None)
        if additional_found < 1:
            return None
        
        extracted_data = {req_col: df[act_col] for req_col, act_col in column_mapping.items() if act_col in df.columns}
        if not extracted_data:
            return None
        
        if 'Loose Part No' in column_mapping and column_mapping['Loose Part No'] in df.columns:
            loose_col = column_mapping['Loose Part No']

            if 'Part No' in extracted_data:
                part_no_series = extracted_data['Part No'].fillna('').astype(str).str.strip()
                loose_series = df[loose_col].fillna('').astype(str).str.strip()

                extracted_data['Part No'] = part_no_series.where(part_no_series != '', loose_series)
            else:
                extracted_data['Part No'] = df[loose_col]

        if not extracted_data:
            return None
        df_result = pd.DataFrame(extracted_data)
        
        if 'Part No' in df_result.columns:
            df_result['Part No'] = df_result['Part No'].astype(str).str.strip()
            df_result['Part No'] = df_result['Part No'].str.replace(r'\.0', '', regex=True)
            df_result['Part No'] = df_result['Part No'].replace(['', 'nan', 'None'], None)
            df_result['Part No'] = df_result['Part No'].ffill()
            df_result = df_result[df_result['Part No'].notna()]
            df_result = df_result[df_result['Part No'] != 'nan']
            df_result = df_result[df_result['Part No'] != '']
            
            df_result = df_result[df_result['Part No'].apply(self.is_valid_part_no) & df_result.apply(self.has_sufficient_data, axis=1)]
        
        df_result = df_result.dropna(how='all')
        
        if len(df_result) == 0:
            return None
        
        rename_map = {k: self.REQUIRED_COLUMNS[k] for k in extracted_data.keys() if k in self.REQUIRED_COLUMNS}
        df_result = df_result.rename(columns=rename_map)
        
        if 'Validity' in df_result.columns:
            df_result['Validity'] = pd.to_numeric(df_result['Validity'], errors='coerce')
        
        if 'PART NO' in df_result.columns:
            df_result['PART NO'] = df_result['PART NO'].astype(str)
        
        df_result.insert(0, 'LEVEL', '')
        type_position = df_result.columns.get_loc('PART NAME') + 1 if 'PART NAME' in df_result.columns else 3
        df_result.insert(type_position, 'Type', '')
        commodity_position = df_result.columns.get_loc('MOD') + 1 if 'MOD' in df_result.columns else 5
        df_result.insert(commodity_position, 'Commodity', '')
        
        return df_result
    
    def process_file(self, uploaded_file):
        """Process a single uploaded file"""
        file_bytes = uploaded_file.read()
        file_name = uploaded_file.name
        file_ext = os.path.splitext(file_name)[1].lower()
        
        try:
            if file_ext == '.xls':
                wb_xls = xlrd.open_workbook(file_contents=file_bytes)
                sheet_names = wb_xls.sheet_names()
            else:
                wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            
            processed_sheets = {}
            
            for sheet_name in sheet_names:
                try:
                    header_row = self.find_header_row(file_bytes, sheet_name, file_ext)
                    engine = 'xlrd' if file_ext == '.xls' else 'openpyxl'
                    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, skiprows=header_row, engine=engine)
                    
                    if len(df) < 1 or len(df.columns) < 3:
                        continue
                    
                    df.columns = [str(col).strip() if not str(col).startswith('Unnamed') else '' for col in df.columns]
                    df = df.loc[:, df.columns != '']
                    df = df.dropna(how='all')
                    
                    if len(df) < 1:
                        continue
                    
                    result_df = self.process_sheet(df, sheet_name)
                    
                    if result_df is not None:
                        processed_sheets[sheet_name] = result_df
                
                except:
                    continue
            
            if not processed_sheets:
                return {'status': 'skipped', 'filename': file_name, 'reason': 'No valid data found'}
            
            # Save to Excel with yellow highlighting
            output_buffer = BytesIO()
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                if len(processed_sheets) == 1:
                    sheet_name = list(processed_sheets.keys())[0]
                    df = processed_sheets[sheet_name]
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    worksheet = writer.sheets['Sheet1']
                    
                    for idx, col in enumerate(df.columns, start=1):
                        if col in ['LEVEL', 'Type', 'Commodity']:
                            for row in range(1, len(df) + 2):
                                worksheet.cell(row=row, column=idx).fill = yellow_fill
                else:
                    for sheet_name, df in processed_sheets.items():
                        safe_sheet_name = sheet_name[:31]
                        df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                        worksheet = writer.sheets[safe_sheet_name]
                        
                        for idx, col in enumerate(df.columns, start=1):
                            if col in ['LEVEL', 'Type', 'Commodity']:
                                for row in range(1, len(df) + 2):
                                    worksheet.cell(row=row, column=idx).fill = yellow_fill
            
            output_buffer.seek(0)
            
            return {
                'status': 'success',
                'filename': file_name,
                'output_filename': os.path.splitext(file_name)[0] + "_processed.xlsx",
                'buffer': output_buffer
            }
        
        except Exception as e:
            return {'status': 'error', 'filename': file_name, 'reason': str(e)}


# ============================================================================
# STREAMLIT UI
# ============================================================================

def main():
    st.title("📊 Quote Processor")
    st.markdown("---")
    
    # Custom CSS for separated column-style tabs
    st.markdown("""
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 100px;
            justify-content: center;
            background-color: transparent;
            padding: 20px 0px;
        }
        .stTabs [data-baseweb="tab"] {
            height: 80px;
            min-width: 300px;
            padding: 0px 40px;
            background-color: #f8f9fa;
            border: 3px solid #dee2e6;
            border-radius: 15px;
            font-weight: 700;
            font-size: 18px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .stTabs [aria-selected="true"] {
            background-color: #0066cc !important;
            color: white !important;
            border-color: #0052a3 !important;
            box-shadow: 0 6px 12px rgba(0,102,204,0.3);
            transform: translateY(-2px);
        }
        </style>
    """, unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["📐 VERTICAL QUOTES", "🤖 HORIZONTAL QUOTES (AI)"])
    
    # ========== VERTICAL QUOTES TAB ==========
    with tab1:
        st.header("📐 Vertical Quote Processing")
        st.markdown("---")
        
        st.info("💡 **Quick Upload Tip:** Navigate to your folder, select all files (Ctrl+A / Cmd+A), and upload them together!")
        
        st.markdown("### 📤 Upload Files")
        uploaded_files_v = st.file_uploader(
            "Select Multiple Vertical Quote Files", 
            type=['xlsx', 'xls'], 
            accept_multiple_files=True, 
            key="vertical",
            help="Select all files from your folder using Ctrl+A (Windows) or Cmd+A (Mac)"
        )
        
        if uploaded_files_v:
            st.markdown("---")
            st.subheader(f"📁 {len(uploaded_files_v)} file(s) ready for processing")
            
            if st.button("🚀 Process Vertical Quotes", type="primary", key="process_v", use_container_width=True):
                processor = VerticalQuoteProcessor()
                
                with st.spinner("Processing..."):
                    successful = []
                    skipped = []
                    errors = []
                    progress = st.progress(0)
                    
                    for i, file in enumerate(uploaded_files_v):
                        result = processor.process_file(file)
                        if result['status'] == 'success':
                            successful.append(result)
                        elif result['status'] == 'skipped':
                            skipped.append(result)
                        else:
                            errors.append(result)
                        progress.progress((i + 1) / len(uploaded_files_v))
                    
                    progress.empty()
                    
                    # Display Statistics
                    st.markdown("---")
                    st.markdown("### 📊 Processing Results")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("✅ Successfully Processed", len(successful), delta=f"{len(successful)} files")
                    with col2:
                        st.metric("⚠️ Skipped Files", len(skipped))
                    with col3:
                        st.metric("❌ Failed Files", len(errors))
                    st.markdown("---")
                    
                    if successful:
                        st.success(f"✅ Successfully processed {len(successful)} file(s)")
                        
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for result in successful:
                                zip_file.writestr(result['output_filename'], result['buffer'].getvalue())
                        zip_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download All Processed Files (ZIP)",
                            data=zip_buffer,
                            file_name=f"vertical_quotes_processed.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                        
                        with st.expander("📋 View processed files"):
                            for result in successful:
                                st.write(f"✓ {result['filename']} → {result['output_filename']}")
                    
                    if skipped:
                        st.warning(f"⚠️ Skipped {len(skipped)} file(s)")
                        with st.expander("📋 View skipped files and reasons"):
                            for result in skipped:
                                st.write(f"**{result['filename']}**")
                                st.write(f"   ↳ Reason: {result['reason']}")
                                st.write("")
                    
                    if errors:
                        st.error(f"❌ Failed to process {len(errors)} file(s)")
                        with st.expander("📋 View error details"):
                            for result in errors:
                                st.write(f"**{result['filename']}**")
                                st.write(f"   ↳ Error: {result['reason']}")
                                st.write("")

    # ========== HORIZONTAL QUOTES TAB ==========
    with tab2:
        st.header("🤖 Horizontal Quote Processing (AI)")
        st.markdown("---")
        
        st.info("💡 **Quick Upload Tip:** Navigate to your folder, select all files (Ctrl+A / Cmd+A), and upload them together!")
        
        api_key = os.getenv("GEMINI_API_KEY")
    
        if not api_key:
            st.error("⚠️ API Key not found! Please add GEMINI_API_KEY to your .env file")
            st.stop()
        
        st.markdown("### 📤 Upload Files")
        uploaded_files_h = st.file_uploader(
            "Select Multiple Horizontal Quote Files", 
            type=['xlsx', 'xls'], 
            accept_multiple_files=True, 
            key="horizontal",
            help="Select all files from your folder using Ctrl+A (Windows) or Cmd+A (Mac)"
        )
        
        if uploaded_files_h:
            st.markdown("---")
            st.subheader(f"📁 {len(uploaded_files_h)} file(s) ready for AI processing")
            
            if st.button("🤖 Process Horizontal Quotes (AI)", type="primary", key="process_h", use_container_width=True):
                processor = HorizontalQuoteProcessor(api_key)
                
                with st.spinner("AI Processing..."):
                    successful = []
                    skipped = []
                    errors = []
                    progress = st.progress(0)
                    
                    for i, file in enumerate(uploaded_files_h):
                        result = processor.process_file(file)
                        if result['status'] == 'success':
                            successful.append(result)
                        elif result['status'] == 'skipped':
                            skipped.append(result)
                        else:
                            errors.append(result)
                        progress.progress((i + 1) / len(uploaded_files_h))
                    
                    progress.empty()
                    
                    # Display Statistics
                    st.markdown("---")
                    st.markdown("### 📊 AI Processing Results")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("✅ Successfully Processed", len(successful), delta=f"{len(successful)} files")
                    with col2:
                        st.metric("⚠️ Skipped Files", len(skipped))
                    with col3:
                        st.metric("❌ Failed Files", len(errors))
                    st.markdown("---")
                    
                    if successful:
                        st.success(f"✅ Successfully processed {len(successful)} file(s)")
                        
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                            for result in successful:
                                zip_file.writestr(result['output_filename'], result['buffer'].getvalue())
                        zip_buffer.seek(0)
                        
                        st.download_button(
                            label="📥 Download All Processed Files (ZIP)",
                            data=zip_buffer,
                            file_name=f"horizontal_quotes_ai_processed.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                        
                        with st.expander("📋 View processed files"):
                            for result in successful:
                                st.write(f"✓ {result['filename']} → {result['output_filename']}")
                    
                    if skipped:
                        st.warning(f"⚠️ Skipped {len(skipped)} file(s)")
                        with st.expander("📋 View skipped files and reasons"):
                            for result in skipped:
                                st.write(f"**{result['filename']}**")
                                st.write(f"   ↳ Reason: {result['reason']}")
                                st.write("")
                    
                    if errors:
                        st.error(f"❌ Failed to process {len(errors)} file(s)")
                        with st.expander("📋 View error details"):
                            for result in errors:
                                st.write(f"**{result['filename']}**")
                                st.write(f"   ↳ Error: {result['reason']}")
                                st.write("")

if __name__ == "__main__":
    main()
